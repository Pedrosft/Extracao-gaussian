import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import re
import os

def extrair_dados_arquivo(nome_arquivo):
    with open(nome_arquivo, 'r') as arquivo:
        dados = arquivo.read()

    # Expressões regulares para capturar os valores de Q da segunda tabela
    electronic_matches = re.finditer(r'Electronic\s+([\d.]+D[+-]\d+)', dados)
    electronic_values = [(match.group(1)) for match in electronic_matches]

    translational_matches = re.finditer(r'Translational\s+([\d.]+D[+-]\d+)', dados)
    translational_values = [(match.group(1)) for match in translational_matches]

    rotational_matches = re.finditer(r'Rotational\s+([\d.]+D[+-]\d+)', dados)
    rotational_values = [(match.group(1)) for match in rotational_matches]

    # Extrair outras propriedades usando expressões regulares
    temperatura_match = re.search(r'Temperature\s+([\d.]+)\s+Kelvin\.', dados)
    temperatura = float(temperatura_match.group(1)) if temperatura_match else None

    entalpia_match = re.search(r'Sum of electronic and thermal Enthalpies=\s+(-?\d+\.\d+)', dados)
    entalpia = float(entalpia_match.group(1)) if entalpia_match else None

    energia_livre_gibbs_match = re.search(r'Sum of electronic and thermal Free Energies=\s+(-?\d+\.\d+)', dados)
    energia_livre_gibbs = float(energia_livre_gibbs_match.group(1)) if energia_livre_gibbs_match else None

    entropia_match = re.search(r'Total\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)', dados)
    entropia = float(entropia_match.group(3)) if entropia_match else None

    cv_match = re.search(r'Total\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)', dados)
    cv = float(cv_match.group(2)) if cv_match else None

    energia_match = re.search(r'Total\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)', dados)
    energia = float(energia_match.group(1)) if energia_match else None

    pressure_match = re.search(r'Pressure\s+([\d.]+)\s+Atm\.', dados)
    pressure = float(pressure_match.group(1)) if pressure_match else None

    molecular_mass_match = re.search(r'Molecular mass:\s+([\d.]+)\s+amu\.', dados)
    molecular_mass = float(molecular_mass_match.group(1)) if molecular_mass_match else None

    return (temperatura, entalpia, energia_livre_gibbs, entropia, cv, energia, pressure, molecular_mass,
            electronic_values, translational_values, rotational_values)

def adicionar_dados_excel(arquivo_excel, dados):
    # Verifica se o arquivo Excel já existe
    if not os.path.exists(arquivo_excel):
        workbook = openpyxl.Workbook()
        workbook.save(arquivo_excel)

    # Carrega o arquivo Excel existente
    workbook = openpyxl.load_workbook(arquivo_excel)
    sheet = workbook.active

    # Insere os cabeçalhos na primeira linha, se necessário
    if sheet['A1'].value is None:
        sheet['A1'] = 'Molécula'
        sheet['B1'] = 'Temperatura (K)'
        sheet['C1'] = 'Entalpia'
        sheet['D1'] = 'Energia Livre de Gibbs'
        sheet['E1'] = 'Entropia'
        sheet['F1'] = 'CV'
        sheet['G1'] = 'E (Thermal)'
        sheet['H1'] = 'Pressão (Atm)'
        sheet['I1'] = 'Massa molecular (amu)'
        sheet['J1'] = 'Electronic (Q)'
        sheet['K1'] = 'Translational (Q)'
        sheet['L1'] = 'Rotational (Q)'

    # Itera sobre os dados para inserir na planilha
    for nome_arquivo, *valores in dados:
        # Extrair apenas o nome do arquivo sem o caminho e a extensão
        nome_arquivo = os.path.basename(nome_arquivo).replace('.LOG', '')

        # Encontrar a próxima linha vazia na planilha
        row = sheet.max_row + 1

        # Insere os dados na próxima linha vazia da planilha
        sheet.cell(row=row, column=1).value = nome_arquivo
        sheet.cell(row=row, column=2).value = valores[0]  # Temperatura
        sheet.cell(row=row, column=3).value = valores[1]  # Entalpia
        sheet.cell(row=row, column=4).value = valores[2]  # Energia Livre de Gibbs
        sheet.cell(row=row, column=5).value = valores[3]  # Entropia
        sheet.cell(row=row, column=6).value = valores[4]  # CV
        sheet.cell(row=row, column=7).value = valores[5]  # Energia
        sheet.cell(row=row, column=8).value = valores[6]  # Pressure
        sheet.cell(row=row, column=9).value = valores[7]  # Molecular Mass
        sheet.cell(row=row, column=10).value = valores[8][0] if len(valores) > 8 and valores[8] else None  # Electronic Q
        sheet.cell(row=row, column=11).value = valores[9][0] if len(valores) > 9 and valores[9] else None  # Translational Q
        sheet.cell(row=row, column=12).value = valores[10][0] if len(valores) > 10 and valores[10] else None  # Rotational Q

    # Salva o arquivo Excel com os novos dados
    workbook.save(arquivo_excel)

def selecionar_arquivos():
    global app
    arquivos_nomes = filedialog.askopenfilenames(filetypes=[("Arquivos de Log", "*.LOG")])
    if arquivos_nomes:
        app.entry_arquivo.delete(0, tk.END)
        app.entry_arquivo.insert(0, ", ".join(arquivos_nomes))

def extrair_dados_e_inserir():
    arquivos_nomes = app.entry_arquivo.get().split(", ")

    if not arquivos_nomes:
        messagebox.showerror("Erro", "Selecione um ou mais arquivos de log.")
        return

    dados = []

    for arquivo_nome in arquivos_nomes:
        dados_extraidos = extrair_dados_arquivo(arquivo_nome)
        if dados_extraidos:
            dados.append((arquivo_nome,) + dados_extraidos)
        else:
            messagebox.showwarning("Aviso", f"Não foi possível extrair os dados de '{arquivo_nome}'. Verifique o formato do arquivo.")

    if dados:
        nome_arquivo_excel = app.entry_nome_arquivo.get().strip()
        if not nome_arquivo_excel:
            messagebox.showerror("Erro", "Digite um nome para o arquivo Excel.")
            return
        
        # Caminho para salvar o arquivo Excel na área de trabalho
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        arquivo_excel = os.path.join(desktop_path, f'{nome_arquivo_excel}.xlsx')

        adicionar_dados_excel(arquivo_excel, dados)
        messagebox.showinfo("Sucesso", f"Dados salvos em '{arquivo_excel}'.")

class App:
    def __init__(self, root):
        self.root = root
        self.frame = None
        self.entry_arquivo = None
        self.entry_nome_arquivo = None
        self.create_login_screen()

    def create_login_screen(self):
        # Personalize o nome do programa abaixo
        self.root.title("CapiQuím | Extração de Dados")
        self.root.geometry("400x400")
        self.root.configure(bg='#ffffff')  # Cor de fundo branca

        # Centralizando o conteúdo
        frame = tk.Frame(self.root, bg='#ffffff')
        frame.pack(expand=True)

        # Botão "Iniciar Programa"
        start_button = tk.Button(frame, text="Iniciar Programa", command=self.create_main_interface, bg='#3498db', fg='#ffffff', font=('Helvetica', 14, 'bold'), bd=0, width=20)
        start_button.pack(pady=20)

        # Label para mostrar a pergunta
        self.question_label = tk.Label(frame, text="Extração de Dados | Guassian", bg='#ffffff', fg='#2c3e50', font=('Helvetica', 12), pady=5)
        self.question_label.pack(side=tk.BOTTOM, pady=(0, 10))

        # Colocando o nome do desenvolvedor no rodapé em tamanho pequeno
        developer_label = tk.Label(self.root, text="Desenvolvido por Pedroos_ft", bg='#ffffff', fg='#2c3e50', font=('Helvetica', 6))
        developer_label.pack(side=tk.BOTTOM, pady=(0, 5))
        developer_label = tk.Label(self.root, text="CapiQuím", bg='#ffffff', fg='#2c3e50', font=('Helvetica', 6))
        developer_label.pack(side=tk.BOTTOM, pady=(0, 5))

    def create_main_interface(self):
        if self.frame:
            self.frame.destroy()

        # Criar a janela principal
        self.root.title("CapiQuím | Extração de Dados")
        self.root.geometry("400x400")  # Tamanho padrão da janela
        self.root.configure(bg="white")  # Fundo branco

        # Criar os widgets da interface centralizados
        self.frame = tk.Frame(self.root, bg="white")
        self.frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        label_arquivo = tk.Label(self.frame, text="Moléculas:", bg="white", fg="black")
        label_arquivo.pack(pady=(10, 0))

        self.entry_arquivo = tk.Entry(self.frame, width=50)
        self.entry_arquivo.pack(pady=(0, 10))

        button_selecionar = tk.Button(self.frame, text="Selecionar Moléculas", bg="lightblue", fg="black", 
        command=selecionar_arquivos)
        button_selecionar.pack(pady=(0, 20))

        label_nome_arquivo = tk.Label(self.frame, text="Nome do Arquivo Excel:", bg="white", fg="black")
        label_nome_arquivo.pack()

        self.entry_nome_arquivo = tk.Entry(self.frame, width=50)
        self.entry_nome_arquivo.pack(pady=(0, 10))

        button_extrair_inserir = tk.Button(self.frame, text="Extrair Dados", bg="lightblue", fg="black", 
        command=extrair_dados_e_inserir)
        button_extrair_inserir.pack(pady=(10, 20))

# Criar a janela principal
root = tk.Tk()
app = App(root)
root.mainloop()
 